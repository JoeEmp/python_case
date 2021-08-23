# 不做研究只做应用，用到哪写到哪

# import xlrd
from datetime import datetime
from openpyxl import load_workbook

datetime_fmt = '%H:%M:%S'


class my_excel():
    def __init__(self, xls, sheet_name='Sheet1'):
        self.xls, self.sheet_name = xls, sheet_name
        self.header, self.rows = self.read_excel_xls()
        self.to_dict()

    def read_excel_xls(self, xls='', sheet_name=''):
        """读取excel,返回字段名和数据(以行为单位) 
        [['name','local'],
         ['joe','china']]
        """
        xls = xls or self.xls
        sheet_name = sheet_name or self.sheet_name
        self.workbook = load_workbook(xls)
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        rows = []
        for row in range(1, self.sheet.max_row+1):
            line = []
            for col in range(1, self.sheet.max_column+1):
                line.append(self.dispaly_value(row, col))
            rows.append(line)
        header = rows.pop(0)
        return header, rows

    def dispaly_value(self, row, col):
        """return dispaly value which your edit"""
        # 兼容日期value为浮点数的情况
        return self.sheet.cell(row=row, column=col).value

    def to_dict(self, header=None, rows=None):
        """行数据转json格式
        [{"name","joe"},{"local","china"}]
        """
        self.dict_table = []
        rows = rows or self.rows
        header = header or self.header
        for row in rows:
            d = {}
            for i in range(len(row)):
                key, value = header[i], row[i]
                d[key] = value
            self.dict_table.append(d)

    def write_cell(self, value, row, col):
        self.sheet.cell(row=row, column=col, value=value)

    def save(self):
        self.workbook.save(self.xls)


def main():
    excel = my_excel('/Users/joe/Documents/eos埋点.xlsx', 'point')
    lines = excel.dict_table
    for i in range(len(lines)):
        if not lines[i]['备注(extra_field)']:
            lines[i]['query'] = '"{}" and "{}" and "{}"'.format(
                lines[i]['页面（page）'],
                lines[i]['事件类型（event_code）'],
                lines[i]['事件名称（specific_event_key）']
            )
        else:
            lines[i]['query'] = ''
            for param in lines[i]['备注(extra_field)'].split('\n'):
                lines[i]['query'] += '"{}" and "{}" and "{}" and "{}"'.format(
                    lines[i]['页面（page）'],
                    lines[i]['事件类型（event_code）'],
                    lines[i]['事件名称（specific_event_key）'],
                    param
                ) + '\n'
        excel.write_cell(lines[i]['query'], i+2, 1)
    excel.save()


main()
